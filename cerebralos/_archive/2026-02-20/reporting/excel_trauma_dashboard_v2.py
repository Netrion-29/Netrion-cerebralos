#!/usr/bin/env python3
"""
CerebralOS Trauma Excellence Dashboard v2 — Elle Woods Theme.

Auto-generates an Excel workbook from pipeline outputs (green card, features,
timeline, PI reports) with 7 sheets.  Manual columns are NEVER overwritten
on re-run, so Sarah can add room numbers, notes, DC info at work.

Output: outputs/trauma_excellence_dashboard.xlsx (.xlsx, no macros)

Usage:
    python -m cerebralos.reporting.excel_trauma_dashboard_v2 --patient Dallas_Clark
    python -m cerebralos.reporting.excel_trauma_dashboard_v2 --all
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
_OUTPUTS = _PROJECT_ROOT / "outputs"
_DEFAULT_OUTPUT = _OUTPUTS / "trauma_excellence_dashboard.xlsx"


# ---------------------------------------------------------------------------
# Elle Woods color palette (reused from excel_dashboard.py)
# ---------------------------------------------------------------------------
if HAS_OPENPYXL:
    _PINK_FILL = PatternFill(start_color="FFEC4899", end_color="FFEC4899", fill_type="solid")
    _PINK_LIGHT = PatternFill(start_color="FFFDF2F8", end_color="FFFDF2F8", fill_type="solid")
    _GOLD_FILL = PatternFill(start_color="FFD4A843", end_color="FFD4A843", fill_type="solid")
    _GOLD_LIGHT = PatternFill(start_color="FFFEF3C7", end_color="FFFEF3C7", fill_type="solid")
    _EMERALD_FILL = PatternFill(start_color="FF059669", end_color="FF059669", fill_type="solid")
    _EMERALD_LIGHT = PatternFill(start_color="FFD1FAE5", end_color="FFD1FAE5", fill_type="solid")
    _GRAY_FILL = PatternFill(start_color="FFF3F4F6", end_color="FFF3F4F6", fill_type="solid")
    _WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    _HEADER_FONT = Font(name="Georgia", size=11, bold=True, color="FFFFFFFF")
    _HEADER_FILL = PatternFill(start_color="FFBE185D", end_color="FFBE185D", fill_type="solid")
    _BODY_FONT = Font(name="Calibri", size=10)
    _BOLD_FONT = Font(name="Calibri", size=10, bold=True)
    _PINK_FONT = Font(name="Calibri", size=10, bold=True, color="FFBE185D")
    _GOLD_FONT = Font(name="Calibri", size=10, bold=True, color="FFB8860B")
    _EMERALD_FONT = Font(name="Calibri", size=10, bold=True, color="FF059669")
    _THIN_BORDER = Border(
        left=Side(style="thin", color="FFFCE7F3"),
        right=Side(style="thin", color="FFFCE7F3"),
        top=Side(style="thin", color="FFFCE7F3"),
        bottom=Side(style="thin", color="FFFCE7F3"),
    )
else:
    _PINK_FILL = _PINK_LIGHT = _GOLD_FILL = _GOLD_LIGHT = None
    _EMERALD_FILL = _EMERALD_LIGHT = _GRAY_FILL = _WHITE_FILL = None
    _HEADER_FONT = _HEADER_FILL = _BODY_FONT = _BOLD_FONT = None
    _PINK_FONT = _GOLD_FONT = _EMERALD_FONT = _THIN_BORDER = None


# ---------------------------------------------------------------------------
# Manual column sets (1-indexed) — NEVER overwritten on existing rows
# ---------------------------------------------------------------------------
_MANUAL_COLS_TRACKER = {27, 28, 29, 30, 31}       # Sheet 1: AA-AE
_MANUAL_COLS_LABS = {20}                            # Sheet 2: T
_MANUAL_COLS_PROTOCOL = set()                       # dynamic — last col
_MANUAL_COLS_NTDS = set()                           # dynamic — last col
_MANUAL_COLS_IMAGING = {8, 9, 10, 11}              # Sheet 5: H-K
_MANUAL_COLS_SUMMARY = {9, 10}                      # Sheet 6: I-J
_MANUAL_COLS_CONSULTS = {8, 9}                      # Sheet 7: H-I


# ---------------------------------------------------------------------------
# Lab mapping: features daily key -> display name
# ---------------------------------------------------------------------------
_LAB_MAP = {
    "Hemoglobin":           ("Hgb",           5),
    "Platelet Count":       ("Plt",           7),
    "White Blood Cell Count": ("WBC",         9),
    "INR":                  ("INR",          11),
    "Lactate":              ("Lactate",      12),
    "Base Deficit":         ("Base Deficit", 13),
    "Creatinine":           ("Creatinine",   14),
    "Blood Urea Nitrogen":  ("BUN",          15),
    "Sodium":               ("Na",           16),
    "Potassium":            ("K",            17),
    "Glucose":              ("Glucose",      18),
}

# Trend columns (value col + 1)
_TREND_LABS = {"Hemoglobin": 6, "Platelet Count": 8, "White Blood Cell Count": 10}


# ---------------------------------------------------------------------------
# NTDS canonical event IDs (in standard order)
# ---------------------------------------------------------------------------
_NTDS_EVENT_IDS = [
    1, 2, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21,
]


# ---------------------------------------------------------------------------
# Helper: style header row
# ---------------------------------------------------------------------------
def _style_header_row(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


# ---------------------------------------------------------------------------
# Helper: find rows by key columns
# ---------------------------------------------------------------------------
def _find_patient_row(ws, mrn: str) -> Optional[int]:
    """Find existing row by MRN in column B."""
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=2).value or "") == str(mrn):
            return row
    return None


def _find_patient_day_row(ws, mrn: str, dt: str) -> Optional[int]:
    """Find existing row by MRN (col B) + Date (col C)."""
    for row in range(2, ws.max_row + 1):
        if (str(ws.cell(row=row, column=2).value or "") == str(mrn)
                and str(ws.cell(row=row, column=3).value or "") == str(dt)):
            return row
    return None


def _find_item_row(ws, mrn: str, source_id: str) -> Optional[int]:
    """Find existing row by MRN (col B) + Source ID (col D)."""
    for row in range(2, ws.max_row + 1):
        if (str(ws.cell(row=row, column=2).value or "") == str(mrn)
                and str(ws.cell(row=row, column=4).value or "") == str(source_id)):
            return row
    return None


def _next_row(ws) -> int:
    """Return the next empty row (handles header-only sheets)."""
    mx = ws.max_row
    if mx == 1 and ws.cell(row=2, column=1).value is None:
        return 2
    return mx + 1


def _get_or_append_row(ws, finder, manual_cols: Set[int]) -> Tuple[int, bool]:
    """Returns (row_number, is_new).  finder returns Optional[int]."""
    existing = finder()
    if existing:
        return existing, False
    return _next_row(ws), True


# ---------------------------------------------------------------------------
# Helper: write a row of data, preserving manual columns
# ---------------------------------------------------------------------------
def _write_row(ws, row: int, data: Dict[int, Any], manual_cols: Set[int],
               is_new: bool) -> None:
    """Write data dict {col: value} to a row, skipping manual cols on existing rows."""
    for col, val in data.items():
        if col in manual_cols and not is_new:
            continue  # preserve manual data
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _BODY_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(col > 20))

    # Initialize manual cols to empty on new rows
    if is_new:
        for col in manual_cols:
            if col not in data:
                cell = ws.cell(row=row, column=col, value="")
                cell.font = _BODY_FONT
                cell.border = _THIN_BORDER


# ---------------------------------------------------------------------------
# Helper: conditional formatting
# ---------------------------------------------------------------------------
def _add_conditional_formatting(ws, col_letter: str, max_row: int,
                                values_colors: List) -> None:
    for val, fill in values_colors:
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{max_row}",
            CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill),
        )


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def load_patient_data(patient_slug: str) -> Dict[str, Any]:
    """Load all pipeline JSON outputs for a patient.  Returns dict with keys:
    green_card, features, timeline, pi_results.  Missing files → None.
    """
    data: Dict[str, Any] = {"slug": patient_slug}

    gc_path = _OUTPUTS / "green_card" / patient_slug / "green_card_v1.json"
    if gc_path.exists():
        data["green_card"] = json.loads(gc_path.read_text(encoding="utf-8"))
    else:
        data["green_card"] = None

    feat_path = _OUTPUTS / "features" / patient_slug / "patient_features_v1.json"
    if feat_path.exists():
        data["features"] = json.loads(feat_path.read_text(encoding="utf-8"))
    else:
        data["features"] = None

    tl_path = _OUTPUTS / "timeline" / patient_slug / "patient_days_v1.json"
    if tl_path.exists():
        data["timeline"] = json.loads(tl_path.read_text(encoding="utf-8"))
    else:
        data["timeline"] = None

    # PI results — try underscore name first, then space name
    pi_path = _OUTPUTS / "pi_reports" / f"{patient_slug}_results.json"
    if not pi_path.exists():
        # Try with spaces instead of underscores
        space_name = patient_slug.replace("_", " ")
        pi_path = _OUTPUTS / "pi_reports" / f"{space_name}_results.json"
    if pi_path.exists():
        data["pi_results"] = json.loads(pi_path.read_text(encoding="utf-8"))
    else:
        data["pi_results"] = None

    return data


def _safe_get(obj: Any, *keys: str, default: Any = "") -> Any:
    """Safely navigate nested dicts."""
    cur = obj
    for k in keys:
        if isinstance(cur, dict):
            cur = cur.get(k)
        else:
            return default
        if cur is None:
            return default
    return cur


def _compute_age(dob_str: str, ref_date: Optional[str] = None) -> Optional[int]:
    """Compute age from DOB string (M/D/YYYY) relative to ref_date or today."""
    if not dob_str:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            dob = datetime.strptime(dob_str, fmt).date()
            ref = date.today()
            if ref_date:
                try:
                    ref = datetime.strptime(ref_date[:10], "%Y-%m-%d").date()
                except ValueError:
                    pass
            age = ref.year - dob.year - ((ref.month, ref.day) < (dob.month, dob.day))
            return age
        except ValueError:
            continue
    return None


def _parse_date(dt_str: str) -> str:
    """Extract YYYY-MM-DD from a datetime string."""
    if not dt_str:
        return ""
    return dt_str[:10]


def _parse_time(dt_str: str) -> str:
    """Extract HH:MM from a datetime string."""
    if not dt_str:
        return ""
    if len(dt_str) >= 16:
        return dt_str[11:16]
    return ""


def _yn(val: Any) -> str:
    """Convert truthy/falsy to Y/N."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Y" if val else "N"
    s = str(val).lower().strip()
    if s in ("true", "yes", "y", "1"):
        return "Y"
    if s in ("false", "no", "n", "0"):
        return "N"
    return str(val)


def _has_head_injury(gc: Dict) -> str:
    """Determine if patient has head injury from green card data."""
    injuries = _safe_get(gc, "injuries", "value", default=[])
    if isinstance(injuries, list):
        for inj in injuries:
            if isinstance(inj, str) and re.search(r"head|sah|sdh|edh|tbi|brain|intracranial", inj, re.I):
                return "Y"
    moi = _safe_get(gc, "moi_narrative", "value", default="")
    hpi = _safe_get(gc, "hpi", "value", default="")
    combined = f"{moi} {hpi}"
    if re.search(r"subarachnoid|subdural|epidural|intracranial|head injury|head trauma|tbi|brain", combined, re.I):
        return "Y"
    imp = _safe_get(gc, "impression_plan", "entries", default=[])
    if isinstance(imp, list):
        for entry in imp:
            if isinstance(entry, dict):
                text = entry.get("impression", "") or entry.get("text", "")
            else:
                text = str(entry)
            if re.search(r"SAH|SDH|EDH|TBI|subarachnoid|subdural|head", text, re.I):
                return "Y"
    return "N"


def _truncate(text: str, max_len: int = 200) -> str:
    """Truncate text to max_len characters."""
    if not text:
        return ""
    text = text.replace("\n", " ").strip()
    if len(text) <= max_len:
        return text
    return text[:max_len - 3] + "..."


def _extract_impression(payload_text: str) -> str:
    """Extract IMPRESSION section from radiology text."""
    if not payload_text:
        return ""
    m = re.search(r"IMPRESSION:\s*(.*?)(?:\n\n|\Z)", payload_text, re.DOTALL)
    if m:
        return _truncate(m.group(1).strip(), 300)
    return ""


def _extract_study_type(payload_text: str) -> str:
    """Extract study type from radiology payload."""
    if not payload_text:
        return ""
    # First line after [RADIOLOGY] tag is usually the study type
    lines = payload_text.split("\n")
    for line in lines:
        line = line.strip()
        if line.startswith("[RADIOLOGY]"):
            continue
        if line.startswith("Result Date:"):
            continue
        if line and not line.startswith("["):
            return _truncate(line, 80)
    return ""


def _extract_body_region(study_type: str) -> str:
    """Infer body region from study type."""
    st = study_type.upper()
    for region, patterns in [
        ("Head", ["HEAD", "BRAIN"]),
        ("C-Spine", ["CERVICAL", "C-SPINE", "CSPINE"]),
        ("T-Spine", ["THORACIC SPINE", "T-SPINE"]),
        ("L-Spine", ["LUMBAR", "L-SPINE"]),
        ("Chest", ["CHEST", "THORAX", "LUNG"]),
        ("Abdomen/Pelvis", ["ABDOMEN", "PELVIS", "ABD"]),
        ("Extremity", ["FEMUR", "TIBIA", "HIP", "SHOULDER", "HUMERUS", "KNEE", "ANKLE", "WRIST", "FOREARM"]),
        ("Face", ["FACE", "FACIAL", "ORBIT", "MAXILLO"]),
    ]:
        if any(p in st for p in patterns):
            return region
    return ""


def _extract_consult_service(payload_text: str) -> str:
    """Extract consulting service from a CONSULT_NOTE payload."""
    if not payload_text:
        return ""
    text_lower = payload_text.lower()
    services = [
        ("Orthopedics", ["ortho", "orthopedic"]),
        ("Neurosurgery", ["neurosurg", "nsgy"]),
        ("Trauma Surgery", ["trauma surg"]),
        ("Critical Care", ["critical care", "icu consult", "intensivist"]),
        ("Cardiology", ["cardiology", "cardio consult"]),
        ("Neurology", ["neurology"]),
        ("Pulmonology", ["pulmonol"]),
        ("Physical Therapy", ["physical therap", " pt consult", "pt/ot"]),
        ("Occupational Therapy", ["occupational therap"]),
        ("Social Work", ["social work", "case manage"]),
        ("Nutrition", ["nutrition", "dietitian"]),
        ("Pharmacy", ["pharmacy", "pharmacist"]),
        ("Pain Management", ["pain manage"]),
    ]
    for svc, keywords in services:
        if any(kw in text_lower for kw in keywords):
            return svc
    return "Consult"


def _extract_note_kind(payload_text: str) -> str:
    """Extract note kind from consult text."""
    if not payload_text:
        return ""
    if re.search(r"initial consult|reason for consult|new consult", payload_text, re.I):
        return "Initial"
    if re.search(r"follow.?up|f/u|subsequent", payload_text, re.I):
        return "Follow-up"
    if re.search(r"progress note", payload_text, re.I):
        return "Progress"
    return "Note"


# ---------------------------------------------------------------------------
# Workbook creation
# ---------------------------------------------------------------------------
def _create_workbook() -> "Workbook":
    """Create a new workbook with all 7 sheets, headers, widths, formatting."""
    wb = Workbook()

    # ── Sheet 1: Patient Tracker ──
    ws1 = wb.active
    ws1.title = "Patient Tracker"
    s1_headers = [
        "Name", "MRN", "DOB", "Age", "Admit Date", "Admit Time",
        "LOS", "MOI", "GCS", "Trauma Cat", "Attending MD",
        "Transfer From", "Head Injury", "Admit Service", "Code Status",
        "Spine Clear", "DVT Ppx Agent", "DVT Ppx Started",
        "ETOH", "UDS", "Base Deficit", "INR",
        "First ED Temp", "Injuries", "PMH", "Home Anticoags",
        # Manual columns (AA-AE)
        "Room", "Status Notes", "DC Date", "DC Destination", "Reviewer Notes",
    ]
    for col, h in enumerate(s1_headers, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header_row(ws1, len(s1_headers))
    ws1.freeze_panes = "C2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(s1_headers))}1"
    s1_widths = [
        22, 12, 12, 6, 12, 8,   # A-F
        6, 18, 5, 5, 20,        # G-K
        18, 5, 12, 12,          # L-O
        10, 14, 14,             # P-R
        8, 8, 10, 6,            # S-V
        10, 40, 40, 25,         # W-Z
        8, 25, 12, 18, 30,      # AA-AE
    ]
    for i, w in enumerate(s1_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # DC Destination dropdown
    dv_dc = DataValidation(
        type="list",
        formula1='"Home,SNF,Rehab,LTACH,AMA,Expired,Transfer,Hospice"',
        allow_blank=True,
    )
    dv_dc.prompt = "Select discharge destination"
    dv_dc.promptTitle = "DC Destination"
    ws1.add_data_validation(dv_dc)
    dv_dc.add("AD2:AD500")

    # ── Sheet 2: Daily Labs ──
    ws2 = wb.create_sheet("Daily Labs")
    s2_headers = [
        "Name", "MRN", "Date", "Hospital Day#",
        "Hgb", "Hgb Trend", "Plt", "Plt Trend", "WBC", "WBC Trend",
        "INR", "Lactate", "Base Deficit", "Creatinine", "BUN",
        "Na", "K", "Glucose", "Alert Flag",
        # Manual
        "Lab Notes",
    ]
    for col, h in enumerate(s2_headers, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header_row(ws2, len(s2_headers))
    ws2.freeze_panes = "C2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(s2_headers))}1"
    s2_widths = [
        22, 12, 12, 6,                # A-D
        8, 10, 8, 10, 8, 10,          # E-J
        6, 8, 10, 10, 8,              # K-O
        6, 6, 8, 12,                  # P-S
        30,                           # T
    ]
    for i, w in enumerate(s2_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Protocol Tracking ──
    ws3 = wb.create_sheet("Protocol Tracking")
    # Headers for protocols are built dynamically, but we set Name + MRN
    ws3.cell(row=1, column=1, value="Name").font = _HEADER_FONT
    ws3.cell(row=1, column=1).fill = _HEADER_FILL
    ws3.cell(row=1, column=2, value="MRN").font = _HEADER_FONT
    ws3.cell(row=1, column=2).fill = _HEADER_FILL
    ws3.freeze_panes = "C2"
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 12

    # ── Sheet 4: NTDS Events ──
    ws4 = wb.create_sheet("NTDS Events")
    # Fixed headers: Name, MRN, Admit Date, DC Date, then 21 event columns + Notes
    s4_base = ["Name", "MRN", "Admit Date", "DC Date"]
    for col, h in enumerate(s4_base, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    ws4.freeze_panes = "C2"
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 12
    ws4.column_dimensions["D"].width = 12

    # ── Sheet 5: Imaging ──
    ws5 = wb.create_sheet("Imaging")
    s5_headers = [
        "Name", "MRN", "Study Date", "Source ID",
        "Study Type", "Body Region", "Impression",
        # Manual
        "Follow-up Needed?", "Follow-up Date", "Completed?", "Notes",
    ]
    for col, h in enumerate(s5_headers, 1):
        ws5.cell(row=1, column=col, value=h)
    _style_header_row(ws5, len(s5_headers))
    ws5.freeze_panes = "C2"
    ws5.auto_filter.ref = f"A1:{get_column_letter(len(s5_headers))}1"
    # Hide Source ID column
    ws5.column_dimensions["D"].hidden = True
    s5_widths = [22, 12, 12, 10, 30, 16, 60, 14, 12, 10, 30]
    for i, w in enumerate(s5_widths, 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    # Follow-up Needed dropdown
    dv_fu = DataValidation(
        type="list", formula1='"Yes,No,Pending"', allow_blank=True,
    )
    dv_fu.prompt = "Follow-up needed?"
    ws5.add_data_validation(dv_fu)
    dv_fu.add("H2:H1000")

    # Completed dropdown
    dv_comp = DataValidation(
        type="list", formula1='"Yes,No,N/A"', allow_blank=True,
    )
    dv_comp.prompt = "Completed?"
    ws5.add_data_validation(dv_comp)
    dv_comp.add("J2:J1000")

    # ── Sheet 6: Daily Summary ──
    ws6 = wb.create_sheet("Daily Summary")
    s6_headers = [
        "Name", "MRN", "Date", "Hospital Day#",
        "Devices Present", "Services Active", "Key Events", "Procedures",
        # Manual
        "Daily Notes", "Plan",
    ]
    for col, h in enumerate(s6_headers, 1):
        ws6.cell(row=1, column=col, value=h)
    _style_header_row(ws6, len(s6_headers))
    ws6.freeze_panes = "C2"
    ws6.auto_filter.ref = f"A1:{get_column_letter(len(s6_headers))}1"
    s6_widths = [22, 12, 12, 6, 30, 30, 40, 30, 30, 30]
    for i, w in enumerate(s6_widths, 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 7: Consults ──
    ws7 = wb.create_sheet("Consults")
    s7_headers = [
        "Name", "MRN", "Service", "Date", "Source ID",
        "Note Kind", "Text Preview",
        # Manual
        "Response Time", "Notes",
    ]
    for col, h in enumerate(s7_headers, 1):
        ws7.cell(row=1, column=col, value=h)
    _style_header_row(ws7, len(s7_headers))
    ws7.freeze_panes = "C2"
    ws7.auto_filter.ref = f"A1:{get_column_letter(len(s7_headers))}1"
    # Hide Source ID column
    ws7.column_dimensions["E"].hidden = True
    s7_widths = [22, 12, 18, 12, 10, 12, 60, 12, 30]
    for i, w in enumerate(s7_widths, 1):
        ws7.column_dimensions[get_column_letter(i)].width = w

    return wb


# ---------------------------------------------------------------------------
# Sheet 1: Patient Tracker
# ---------------------------------------------------------------------------
def _build_patient_tracker(wb: "Workbook", pdata: Dict) -> None:
    """Populate one patient's row on the Patient Tracker sheet."""
    gc = pdata.get("green_card")
    pi = pdata.get("pi_results")
    feat = pdata.get("features")
    tl = pdata.get("timeline")
    if not gc:
        return

    ws = wb["Patient Tracker"]
    mrn = _safe_get(gc, "meta", "patient_id")
    if not mrn:
        return

    row, is_new = _get_or_append_row(
        ws, lambda: _find_patient_row(ws, mrn), _MANUAL_COLS_TRACKER
    )

    patient_name = _safe_get(pi, "patient_name", default="")
    if not patient_name:
        patient_name = pdata["slug"].replace("_", " ")
    dob = _safe_get(pi, "dob", default="")
    arrival = _safe_get(gc, "meta", "arrival_datetime", default="")
    admit_date = _parse_date(arrival)
    admit_time = _parse_time(arrival)
    age = _compute_age(dob, admit_date)
    moi = _safe_get(gc, "mechanism_of_injury", "value", default="")
    gcs_raw = _safe_get(gc, "primary_survey", "gcs")
    if gcs_raw is None:
        # Try to parse from disability field
        disability = _safe_get(gc, "primary_survey", "disability", default="")
        m = re.search(r"GCS\s*[=:]?\s*(\d+)", str(disability), re.I)
        gcs = m.group(1) if m else ""
    else:
        gcs = str(gcs_raw)
    trauma_cat = _safe_get(gc, "meta", "trauma_category", default="")
    attending = _safe_get(gc, "admitting_md", "name", default="")
    credential = _safe_get(gc, "admitting_md", "credential", default="")
    if credential:
        attending = f"{attending}, {credential}"
    transfer = ""  # derived from HPI if "transfer" mentioned
    hpi = _safe_get(gc, "hpi", "value", default="")
    m_transfer = re.search(r"transfer from\s+(\S+(?:\s+\S+)?)", str(hpi), re.I)
    if m_transfer:
        transfer = m_transfer.group(1).strip().rstrip(".")
    head_injury = _has_head_injury(gc)
    admit_service = _safe_get(gc, "admitting_service", "value", default="")
    code_status = ""  # not currently extracted
    spine_clear = _safe_get(gc, "spine_clearance", "status", default="")
    dvt_agent = _safe_get(gc, "dvt_prophylaxis", "agent", default="")
    dvt_started = _safe_get(gc, "dvt_prophylaxis", "first_admin_dt", default="")
    if dvt_started:
        dvt_started = _parse_date(dvt_started)

    etoh_val = _safe_get(gc, "etoh", "value", default="")
    if etoh_val is None:
        etoh_val = ""
    uds_performed = _safe_get(gc, "uds", "performed", default=False)
    uds_positive = _safe_get(gc, "uds", "positive_flags", default=[])
    if uds_performed and uds_positive:
        uds_str = ", ".join(uds_positive)
    elif uds_performed:
        uds_str = "Negative"
    else:
        uds_str = "Not done"

    base_def = _safe_get(gc, "base_deficit", "value", default="")
    if base_def is None:
        base_def = ""
    inr_val = _safe_get(gc, "inr", "value", default="")
    if inr_val is None:
        inr_val = ""
    first_temp = _safe_get(gc, "first_ed_temp", "value", default="")

    injuries_list = _safe_get(gc, "injuries", "value", default=[])
    if isinstance(injuries_list, list):
        injuries_str = "; ".join(str(x) for x in injuries_list if x)
    else:
        injuries_str = str(injuries_list) if injuries_list else ""

    pmh_list = _safe_get(gc, "pmh", "value", default=[])
    if isinstance(pmh_list, list):
        pmh_str = "; ".join(str(x) for x in pmh_list if x)
    else:
        pmh_str = str(pmh_list) if pmh_list else ""

    anticoags = _safe_get(gc, "home_anticoagulants", "value", default=[])
    if isinstance(anticoags, list):
        anticoags_str = ", ".join(str(x) for x in anticoags if x)
    else:
        anticoags_str = str(anticoags) if anticoags else ""

    data = {
        1: patient_name,
        2: mrn,
        3: dob,
        4: age if age is not None else "",
        5: admit_date,
        6: admit_time,
        7: "",  # LOS — will set formula below
        8: moi,
        9: gcs,
        10: trauma_cat,
        11: attending,
        12: transfer,
        13: head_injury,
        14: admit_service,
        15: code_status,
        16: spine_clear,
        17: dvt_agent,
        18: dvt_started,
        19: etoh_val,
        20: uds_str,
        21: base_def,
        22: inr_val,
        23: first_temp,
        24: _truncate(injuries_str, 300),
        25: _truncate(pmh_str, 300),
        26: anticoags_str,
    }

    _write_row(ws, row, data, _MANUAL_COLS_TRACKER, is_new)

    # LOS formula: =IF(E{row}="","",TODAY()-E{row})
    los_cell = ws.cell(row=row, column=7)
    # If DC Date (col 29) has a value, use that; otherwise use TODAY()
    los_cell.value = f'=IF(E{row}="","",IF(AC{row}<>"",AC{row}-E{row},TODAY()-E{row}))'
    los_cell.number_format = "0"

    # Wrap text for long columns
    for col in (24, 25):
        ws.cell(row=row, column=col).alignment = Alignment(
            vertical="center", wrap_text=True
        )


# ---------------------------------------------------------------------------
# Sheet 2: Daily Labs
# ---------------------------------------------------------------------------
def _build_daily_labs(wb: "Workbook", pdata: Dict) -> None:
    """Populate daily lab rows for one patient."""
    feat = pdata.get("features")
    gc = pdata.get("green_card")
    pi = pdata.get("pi_results")
    if not feat:
        return

    ws = wb["Daily Labs"]
    mrn = feat.get("patient_id", "")
    patient_name = _safe_get(pi, "patient_name",
                             default=pdata["slug"].replace("_", " "))

    # Get arrival date for hospital day calculation
    arrival_date = ""
    if gc:
        arrival_date = _parse_date(_safe_get(gc, "meta", "arrival_datetime", default=""))

    days = feat.get("days", {})
    sorted_dates = sorted(d for d in days.keys() if d != "__UNDATED__")

    for day_idx, dt in enumerate(sorted_dates):
        day_data = days[dt]
        daily_labs = _safe_get(day_data, "labs", "daily", default={})
        if not daily_labs:
            continue

        row, is_new = _get_or_append_row(
            ws, lambda d=dt: _find_patient_day_row(ws, mrn, d), _MANUAL_COLS_LABS
        )

        # Hospital day number
        hosp_day = day_idx + 1

        # Build alert flags
        alerts = []
        for comp_name, comp_data in daily_labs.items():
            if not isinstance(comp_data, dict):
                continue
            if comp_data.get("big_change"):
                alerts.append(f"{comp_name}:BIG_CHANGE")
            if comp_data.get("abnormal_flag_present"):
                alerts.append(f"{comp_name}:ABNORMAL")
        alert_str = "; ".join(alerts) if alerts else ""

        data: Dict[int, Any] = {
            1: patient_name,
            2: mrn,
            3: dt,
            4: hosp_day,
            19: alert_str,
        }

        # Fill lab values
        for comp_name, (display, col) in _LAB_MAP.items():
            comp = daily_labs.get(comp_name)
            if comp and isinstance(comp, dict):
                data[col] = comp.get("last", "")
                # Trend columns
                if comp_name in _TREND_LABS:
                    trend_col = _TREND_LABS[comp_name]
                    delta = comp.get("delta")
                    if delta is not None and delta != 0:
                        direction = "+" if delta > 0 else ""
                        data[trend_col] = f"{direction}{delta:.1f}"
                    else:
                        data[trend_col] = ""

        _write_row(ws, row, data, _MANUAL_COLS_LABS, is_new)

        # Color alert cells
        if alerts:
            alert_cell = ws.cell(row=row, column=19)
            has_big = any("BIG_CHANGE" in a for a in alerts)
            if has_big:
                alert_cell.fill = _PINK_LIGHT
                alert_cell.font = _PINK_FONT
            else:
                alert_cell.fill = _GOLD_LIGHT
                alert_cell.font = _GOLD_FONT

        # Color individual lab cells for abnormals
        for comp_name, (display, col) in _LAB_MAP.items():
            comp = daily_labs.get(comp_name)
            if comp and isinstance(comp, dict):
                cell = ws.cell(row=row, column=col)
                if comp.get("big_change"):
                    cell.fill = _PINK_LIGHT
                elif comp.get("abnormal_flag_present"):
                    cell.fill = _GOLD_LIGHT


# ---------------------------------------------------------------------------
# Sheet 3: Protocol Tracking
# ---------------------------------------------------------------------------
def _build_protocol_tracking(wb: "Workbook", pdata: Dict) -> None:
    """Populate protocol tracking row for one patient."""
    pi = pdata.get("pi_results")
    gc = pdata.get("green_card")
    if not pi:
        return

    ws = wb["Protocol Tracking"]
    mrn = str(pi.get("patient_id", ""))
    patient_name = pi.get("patient_name", pdata["slug"].replace("_", " "))

    row, is_new = _get_or_append_row(
        ws, lambda: _find_patient_row(ws, mrn), set()
    )

    ws.cell(row=row, column=1, value=patient_name).font = _BOLD_FONT
    ws.cell(row=row, column=1).border = _THIN_BORDER
    ws.cell(row=row, column=2, value=mrn).font = _BODY_FONT
    ws.cell(row=row, column=2).border = _THIN_BORDER

    results = pi.get("results", [])
    triggered = [r for r in results if r.get("outcome") != "NOT_TRIGGERED"]

    for i, r in enumerate(triggered):
        col = i + 3
        outcome = r.get("outcome", "")
        short = {
            "COMPLIANT": "C",
            "NON_COMPLIANT": "NC",
            "INDETERMINATE": "I",
            "ERROR": "E",
        }.get(outcome, outcome)

        cell = ws.cell(row=row, column=col, value=short)
        cell.font = _BODY_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

        if outcome == "NON_COMPLIANT":
            cell.fill = _PINK_LIGHT
            cell.font = _PINK_FONT
        elif outcome == "INDETERMINATE":
            cell.fill = _GOLD_LIGHT
            cell.font = _GOLD_FONT
        elif outcome == "COMPLIANT":
            cell.fill = _EMERALD_LIGHT
            cell.font = _EMERALD_FONT

        # Write protocol name in header if not already there
        hdr_cell = ws.cell(row=1, column=col)
        if not hdr_cell.value:
            hdr_cell.value = r.get("protocol_name", "")
            hdr_cell.font = _HEADER_FONT
            hdr_cell.fill = _HEADER_FILL
            hdr_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            hdr_cell.border = _THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = 14

    # DVT Ppx details column (after protocols)
    dvt_col = len(triggered) + 3
    if gc:
        dvt_agent = _safe_get(gc, "dvt_prophylaxis", "agent", default="")
        dvt_dose = _safe_get(gc, "dvt_prophylaxis", "dose", default="")
        dvt_str = f"{dvt_agent} {dvt_dose}".strip() if dvt_agent else ""
        hdr = ws.cell(row=1, column=dvt_col)
        if not hdr.value:
            hdr.value = "DVT Ppx"
            hdr.font = _HEADER_FONT
            hdr.fill = _HEADER_FILL
            hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(dvt_col)].width = 14
        ws.cell(row=row, column=dvt_col, value=dvt_str).font = _BODY_FONT
        ws.cell(row=row, column=dvt_col).border = _THIN_BORDER

    # Rib count column
    rib_col = dvt_col + 1
    # Try to find rib fracture info from injuries
    rib_count = ""
    if gc:
        injuries = _safe_get(gc, "injuries", "value", default=[])
        if isinstance(injuries, list):
            for inj in injuries:
                if isinstance(inj, str) and re.search(r"rib", inj, re.I):
                    m = re.search(r"(\d+)\s*rib", inj, re.I)
                    if m:
                        rib_count = m.group(1)
                    else:
                        rib_count = "Y"
                    break
    hdr = ws.cell(row=1, column=rib_col)
    if not hdr.value:
        hdr.value = "Rib #"
        hdr.font = _HEADER_FONT
        hdr.fill = _HEADER_FILL
        hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(rib_col)].width = 8
    ws.cell(row=row, column=rib_col, value=rib_count).font = _BODY_FONT
    ws.cell(row=row, column=rib_col).border = _THIN_BORDER

    # Protocol Notes (manual, last col)
    notes_col = rib_col + 1
    hdr = ws.cell(row=1, column=notes_col)
    if not hdr.value:
        hdr.value = "Protocol Notes"
        hdr.font = _HEADER_FONT
        hdr.fill = _HEADER_FILL
        hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(notes_col)].width = 30
    # Only set if new row
    notes_cell = ws.cell(row=row, column=notes_col)
    if is_new:
        notes_cell.value = ""
    notes_cell.font = _BODY_FONT
    notes_cell.border = _THIN_BORDER


# ---------------------------------------------------------------------------
# Sheet 4: NTDS Events
# ---------------------------------------------------------------------------
def _build_ntds_events(wb: "Workbook", pdata: Dict) -> None:
    """Populate NTDS events row for one patient."""
    pi = pdata.get("pi_results")
    gc = pdata.get("green_card")
    if not pi:
        return

    ws = wb["NTDS Events"]
    mrn = str(pi.get("patient_id", ""))
    patient_name = pi.get("patient_name", pdata["slug"].replace("_", " "))

    row, is_new = _get_or_append_row(
        ws, lambda: _find_patient_row(ws, mrn), set()
    )

    # Base data
    arrival = _safe_get(gc, "meta", "arrival_datetime", default="") if gc else ""
    admit_date = _parse_date(arrival)
    dc_date = ""
    if pi.get("has_discharge"):
        dc_date = "Discharged"

    ws.cell(row=row, column=1, value=patient_name).font = _BOLD_FONT
    ws.cell(row=row, column=1).border = _THIN_BORDER
    ws.cell(row=row, column=2, value=mrn).font = _BODY_FONT
    ws.cell(row=row, column=2).border = _THIN_BORDER
    ws.cell(row=row, column=3, value=admit_date).font = _BODY_FONT
    ws.cell(row=row, column=3).border = _THIN_BORDER
    ws.cell(row=row, column=4, value=dc_date).font = _BODY_FONT
    ws.cell(row=row, column=4).border = _THIN_BORDER

    ntds = pi.get("ntds_results", [])
    ntds_map = {r.get("event_id"): r for r in ntds}

    for i, eid in enumerate(_NTDS_EVENT_IDS):
        col = i + 5  # offset by Name, MRN, Admit, DC
        r = ntds_map.get(eid)
        if r:
            outcome = r.get("outcome", "")
        else:
            outcome = ""

        cell = ws.cell(row=row, column=col, value=outcome)
        cell.font = _BODY_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

        if outcome == "YES":
            cell.fill = _PINK_LIGHT
            cell.font = _PINK_FONT
        elif outcome in ("UNABLE_TO_DETERMINE", "UNABLE"):
            cell.fill = _GOLD_LIGHT
            cell.font = _GOLD_FONT
        elif outcome == "NO":
            cell.fill = _EMERALD_LIGHT
        elif outcome == "EXCLUDED":
            cell.fill = _GRAY_FILL

        # Write header if not set
        hdr_cell = ws.cell(row=1, column=col)
        if not hdr_cell.value and r:
            canonical = r.get("canonical_name", "")
            hdr_cell.value = f"#{eid:02d} {canonical}"
            hdr_cell.font = _HEADER_FONT
            hdr_cell.fill = _HEADER_FILL
            hdr_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            hdr_cell.border = _THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = 14

    # Event Notes (manual, last col)
    notes_col = 5 + len(_NTDS_EVENT_IDS)
    hdr = ws.cell(row=1, column=notes_col)
    if not hdr.value:
        hdr.value = "Event Notes"
        hdr.font = _HEADER_FONT
        hdr.fill = _HEADER_FILL
        hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(notes_col)].width = 30
    notes_cell = ws.cell(row=row, column=notes_col)
    if is_new:
        notes_cell.value = ""
    notes_cell.font = _BODY_FONT
    notes_cell.border = _THIN_BORDER


# ---------------------------------------------------------------------------
# Sheet 5: Imaging
# ---------------------------------------------------------------------------
def _build_imaging(wb: "Workbook", pdata: Dict) -> None:
    """Populate imaging rows from timeline RADIOLOGY items."""
    tl = pdata.get("timeline")
    gc = pdata.get("green_card")
    pi = pdata.get("pi_results")
    if not tl:
        return

    ws = wb["Imaging"]
    mrn = _safe_get(tl, "meta", "patient_id", default="")
    patient_name = _safe_get(pi, "patient_name",
                             default=pdata["slug"].replace("_", " "))

    seen_source_ids: Set[str] = set()

    for dt, day_obj in sorted(tl.get("days", {}).items()):
        if dt == "__UNDATED__":
            continue
        for item in day_obj.get("items", []):
            if item.get("type") != "RADIOLOGY":
                continue

            source_id = str(item.get("source_id", ""))
            if not source_id or source_id in seen_source_ids:
                continue
            seen_source_ids.add(source_id)

            payload = item.get("payload", {}).get("text", "")
            study_type = _extract_study_type(payload)
            body_region = _extract_body_region(study_type)
            impression = _extract_impression(payload)
            study_date = _parse_date(item.get("dt", ""))

            row, is_new = _get_or_append_row(
                ws,
                lambda sid=source_id: _find_item_row(ws, mrn, sid),
                _MANUAL_COLS_IMAGING,
            )

            data = {
                1: patient_name,
                2: mrn,
                3: study_date,
                4: source_id,
                5: study_type,
                6: body_region,
                7: _truncate(impression, 300),
            }

            _write_row(ws, row, data, _MANUAL_COLS_IMAGING, is_new)

            # Wrap impression
            ws.cell(row=row, column=7).alignment = Alignment(
                vertical="center", wrap_text=True
            )


# ---------------------------------------------------------------------------
# Sheet 6: Daily Summary
# ---------------------------------------------------------------------------
def _build_daily_summary(wb: "Workbook", pdata: Dict) -> None:
    """Populate daily summary rows from features (devices, services)."""
    feat = pdata.get("features")
    tl = pdata.get("timeline")
    gc = pdata.get("green_card")
    pi = pdata.get("pi_results")
    if not feat:
        return

    ws = wb["Daily Summary"]
    mrn = feat.get("patient_id", "")
    patient_name = _safe_get(pi, "patient_name",
                             default=pdata["slug"].replace("_", " "))

    days = feat.get("days", {})
    sorted_dates = sorted(d for d in days.keys() if d != "__UNDATED__")

    for day_idx, dt in enumerate(sorted_dates):
        day_data = days[dt]

        row, is_new = _get_or_append_row(
            ws, lambda d=dt: _find_patient_day_row(ws, mrn, d), _MANUAL_COLS_SUMMARY
        )

        hosp_day = day_idx + 1

        # Devices present
        devices_canonical = _safe_get(day_data, "devices", "canonical", default={})
        present_devices = [
            dev for dev, status in devices_canonical.items()
            if str(status).upper() == "PRESENT"
        ]
        devices_str = ", ".join(present_devices) if present_devices else "None documented"

        # Services active
        service_tags = _safe_get(day_data, "services", "tags", default=[])
        services_str = ", ".join(service_tags) if service_tags else ""

        # Key events from timeline
        key_events = []
        procedures = []
        if tl:
            tl_day = tl.get("days", {}).get(dt, {})
            for item in tl_day.get("items", []):
                itype = item.get("type", "")
                payload = item.get("payload", {}).get("text", "")
                if itype == "RADIOLOGY":
                    study = _extract_study_type(payload)
                    if study:
                        key_events.append(f"Imaging: {study}")
                elif itype in ("TRAUMA_HP", "ED_NOTE"):
                    key_events.append(itype.replace("_", " ").title())
                elif itype == "DISCHARGE":
                    key_events.append("DISCHARGE")

        events_str = "; ".join(key_events[:5]) if key_events else ""
        proc_str = "; ".join(procedures) if procedures else ""

        data = {
            1: patient_name,
            2: mrn,
            3: dt,
            4: hosp_day,
            5: devices_str,
            6: services_str,
            7: _truncate(events_str, 200),
            8: proc_str,
        }

        _write_row(ws, row, data, _MANUAL_COLS_SUMMARY, is_new)


# ---------------------------------------------------------------------------
# Sheet 7: Consults
# ---------------------------------------------------------------------------
def _build_consults(wb: "Workbook", pdata: Dict) -> None:
    """Populate consult rows from timeline CONSULT_NOTE items."""
    tl = pdata.get("timeline")
    pi = pdata.get("pi_results")
    if not tl:
        return

    ws = wb["Consults"]
    mrn = _safe_get(tl, "meta", "patient_id", default="")
    patient_name = _safe_get(pi, "patient_name",
                             default=pdata["slug"].replace("_", " "))

    seen_source_ids: Set[str] = set()

    for dt, day_obj in sorted(tl.get("days", {}).items()):
        if dt == "__UNDATED__":
            continue
        for item in day_obj.get("items", []):
            if item.get("type") != "CONSULT_NOTE":
                continue

            source_id = str(item.get("source_id", ""))
            if not source_id or source_id in seen_source_ids:
                continue
            seen_source_ids.add(source_id)

            payload = item.get("payload", {}).get("text", "")
            service = _extract_consult_service(payload)
            note_kind = _extract_note_kind(payload)
            consult_date = _parse_date(item.get("dt", ""))
            preview = _truncate(payload, 200)

            row, is_new = _get_or_append_row(
                ws,
                lambda sid=source_id: _find_item_row(ws, mrn, sid),
                _MANUAL_COLS_CONSULTS,
            )

            data = {
                1: patient_name,
                2: mrn,
                3: service,
                4: consult_date,
                5: source_id,
                6: note_kind,
                7: preview,
            }

            _write_row(ws, row, data, _MANUAL_COLS_CONSULTS, is_new)

            # Wrap text preview
            ws.cell(row=row, column=7).alignment = Alignment(
                vertical="center", wrap_text=True
            )


# ---------------------------------------------------------------------------
# Conditional formatting pass
# ---------------------------------------------------------------------------
def _apply_conditional_formatting(wb: "Workbook") -> None:
    """Apply conditional formatting across all sheets."""
    max_row = 500  # generous upper bound

    # Sheet 1: Head Injury (col M)
    ws1 = wb["Patient Tracker"]
    _add_conditional_formatting(ws1, "M", max_row, [
        ("Y", _PINK_LIGHT),
        ("N", _EMERALD_LIGHT),
    ])
    # Spine Clear (col P)
    _add_conditional_formatting(ws1, "P", max_row, [
        ("CLEAR", _EMERALD_LIGHT),
        ("NOT_CLEAR", _PINK_LIGHT),
    ])

    # Sheet 3: Protocol outcomes (C/NC/I) — done inline in _build_protocol_tracking

    # Sheet 4: NTDS outcomes — done inline in _build_ntds_events


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def update_trauma_dashboard(
    patient_slug: str,
    output_path: Optional[Path] = None,
) -> Path:
    """
    Add or update a single patient in the Trauma Excellence Dashboard.

    Loads pipeline outputs, opens or creates workbook, populates all 7 sheets.
    Manual columns are preserved on existing rows.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel dashboard generation. "
                          "Install with: pip install openpyxl")

    if output_path is None:
        output_path = _DEFAULT_OUTPUT

    pdata = load_patient_data(patient_slug)
    if not pdata.get("green_card") and not pdata.get("features"):
        print(f"  [SKIP] {patient_slug}: no green_card or features data found")
        return output_path

    # Load or create workbook
    if output_path.exists():
        wb = load_workbook(output_path)
    else:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = _create_workbook()

    # Build all sheets
    _build_patient_tracker(wb, pdata)
    _build_daily_labs(wb, pdata)
    _build_imaging(wb, pdata)
    _build_daily_summary(wb, pdata)
    _build_consults(wb, pdata)
    _build_protocol_tracking(wb, pdata)
    _build_ntds_events(wb, pdata)

    # Apply formatting
    _apply_conditional_formatting(wb)

    wb.save(output_path)
    return output_path


def update_all_patients(output_path: Optional[Path] = None) -> Path:
    """Rebuild the dashboard for all patients with green_card or features data."""
    if output_path is None:
        output_path = _DEFAULT_OUTPUT

    # Collect all patient slugs from green_card and features directories
    slugs: Set[str] = set()
    gc_dir = _OUTPUTS / "green_card"
    if gc_dir.exists():
        for d in gc_dir.iterdir():
            if d.is_dir() and (d / "green_card_v1.json").exists():
                slugs.add(d.name)
    feat_dir = _OUTPUTS / "features"
    if feat_dir.exists():
        for d in feat_dir.iterdir():
            if d.is_dir() and (d / "patient_features_v1.json").exists():
                slugs.add(d.name)

    if not slugs:
        print("No patient data found in outputs/")
        return output_path

    # Remove existing file for clean rebuild
    if output_path.exists():
        output_path.unlink()

    print(f"Building Trauma Excellence Dashboard for {len(slugs)} patients...")
    for slug in sorted(slugs):
        print(f"  Processing: {slug}")
        update_trauma_dashboard(slug, output_path)

    print(f"  Output: {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description="CerebralOS Trauma Excellence Dashboard v2",
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--patient", "-p", help="Patient slug (e.g. Dallas_Clark)")
    group.add_argument("--all", "-a", action="store_true",
                       help="Rebuild for all patients")
    parser.add_argument("--output", "-o", type=Path, default=None,
                        help=f"Output path (default: {_DEFAULT_OUTPUT})")
    args = parser.parse_args()

    output = args.output or _DEFAULT_OUTPUT

    if args.all:
        update_all_patients(output)
    else:
        print(f"Updating dashboard for: {args.patient}")
        result = update_trauma_dashboard(args.patient, output)
        print(f"  Output: {result}")

    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
