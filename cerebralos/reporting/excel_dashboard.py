#!/usr/bin/env python3
"""
CerebralOS Excel Trauma Dashboard Generator — Elle Woods Theme.

Generates an interactive Excel workbook with:
- Patient Dashboard (main overview)
- VRC Injury Categories (detailed matrix)
- Protocol Detail (outcome per protocol per patient)
- NTDS Events (NTDS outcome matrix)

Append-not-overwrite: existing patients are updated, new patients are appended.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Elle Woods color palette (openpyxl uses ARGB hex without #)
# ---------------------------------------------------------------------------
_PINK_FILL = PatternFill(start_color="FFEC4899", end_color="FFEC4899", fill_type="solid") if HAS_OPENPYXL else None
_PINK_LIGHT = PatternFill(start_color="FFFDF2F8", end_color="FFFDF2F8", fill_type="solid") if HAS_OPENPYXL else None
_GOLD_FILL = PatternFill(start_color="FFD4A843", end_color="FFD4A843", fill_type="solid") if HAS_OPENPYXL else None
_GOLD_LIGHT = PatternFill(start_color="FFFEF3C7", end_color="FFFEF3C7", fill_type="solid") if HAS_OPENPYXL else None
_EMERALD_FILL = PatternFill(start_color="FF059669", end_color="FF059669", fill_type="solid") if HAS_OPENPYXL else None
_EMERALD_LIGHT = PatternFill(start_color="FFD1FAE5", end_color="FFD1FAE5", fill_type="solid") if HAS_OPENPYXL else None
_GRAY_FILL = PatternFill(start_color="FFF3F4F6", end_color="FFF3F4F6", fill_type="solid") if HAS_OPENPYXL else None
_WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid") if HAS_OPENPYXL else None

_HEADER_FONT = Font(name="Georgia", size=11, bold=True, color="FFFFFFFF") if HAS_OPENPYXL else None
_HEADER_FILL = PatternFill(start_color="FFBE185D", end_color="FFBE185D", fill_type="solid") if HAS_OPENPYXL else None
_GOLD_FONT = Font(name="Calibri", size=10, bold=True, color="FFB8860B") if HAS_OPENPYXL else None
_BODY_FONT = Font(name="Calibri", size=10) if HAS_OPENPYXL else None
_BOLD_FONT = Font(name="Calibri", size=10, bold=True) if HAS_OPENPYXL else None
_THIN_BORDER = Border(
    left=Side(style="thin", color="FFFCE7F3"),
    right=Side(style="thin", color="FFFCE7F3"),
    top=Side(style="thin", color="FFFCE7F3"),
    bottom=Side(style="thin", color="FFFCE7F3"),
) if HAS_OPENPYXL else None

# VRC category IDs in display order
_VRC_IDS = [
    "NEURO_EPIDURAL_SUBDURAL_TO_OR", "NEURO_SEVERE_TBI_ICU", "NEURO_SPINAL_CORD_DEFICIT",
    "ORTHO_AMPUTATION", "ORTHO_ACETABULAR_PELVIC", "ORTHO_OPEN_FEMUR_TIBIA",
    "ABDTHOR_THORACIC_CARDIAC", "ABDTHOR_SOLID_ORGAN", "ABDTHOR_PENETRATING",
    "NONSURG_ISS9", "NONSURG_GERIATRIC_HIP", "NONSURG_TRANSFER_OUT",
    "ADVERSE_RETURN_SICU_OR", "ADVERSE_ISS25_SURVIVAL",
    "MTP_ACTIVATED", "HOSPICE", "DEATH",
]


def _style_header_row(ws, num_cols: int) -> None:
    """Apply Elle Woods header styling to first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _add_conditional_formatting(ws, col_letter: str, max_row: int, values_colors: List) -> None:
    """Add conditional formatting rules for a column."""
    for val, fill in values_colors:
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{max_row}",
            CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill)
        )


def _find_patient_row(ws, patient_id: str) -> Optional[int]:
    """Find existing row for a patient by ID (column B)."""
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=2).value) == str(patient_id):
            return row
    return None


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_dashboard_sheet(wb: Workbook, evaluation: Dict, vrc_results: List[Dict], row: int) -> None:
    """Populate the Patient Dashboard sheet."""
    ws = wb["Patient Dashboard"]

    results = evaluation.get("results", [])
    ntds = evaluation.get("ntds_results", [])

    nc = sum(1 for r in results if r["outcome"] == "NON_COMPLIANT")
    ind = sum(1 for r in results if r["outcome"] == "INDETERMINATE")
    comp = sum(1 for r in results if r["outcome"] == "COMPLIANT")
    triggered = nc + ind + comp
    ntds_yes_count = sum(1 for r in ntds if r["outcome"] == "YES")
    ntds_yes_names = ", ".join(r["canonical_name"] for r in ntds if r["outcome"] == "YES")
    ntds_unable = sum(1 for r in ntds if r["outcome"] == "UNABLE_TO_DETERMINE")
    vrc_yes = ", ".join(r["category_id"] for r in vrc_results if r["status"] == "YES")
    vrc_possible = ", ".join(r["category_id"] for r in vrc_results if r["status"] == "POSSIBLE")

    # Worst outcome
    if nc > 0:
        worst = "NON_COMPLIANT"
    elif ind > 0:
        worst = "INDETERMINATE"
    elif comp > 0:
        worst = "COMPLIANT"
    else:
        worst = "NOT_TRIGGERED"

    status = "IN HOSPITAL" if evaluation.get("is_live") else "Discharged"
    stem = Path(evaluation.get("source_file", "")).stem

    data = [
        evaluation.get("patient_name", ""),
        evaluation.get("patient_id", ""),
        evaluation.get("dob", ""),
        evaluation.get("arrival_time", ""),
        evaluation.get("trauma_category", ""),
        status,
        triggered,
        comp,
        nc,
        ind,
        ntds_yes_count,
        ntds_yes_names,
        ntds_unable,
        vrc_yes,
        vrc_possible,
        worst,
        f"{stem}_report.html",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]

    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _BODY_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(col in (12, 14, 15)))

    # Color the worst outcome cell
    worst_cell = ws.cell(row=row, column=16)
    if worst == "NON_COMPLIANT":
        worst_cell.fill = _PINK_LIGHT
    elif worst == "INDETERMINATE":
        worst_cell.fill = _GOLD_LIGHT
    elif worst == "COMPLIANT":
        worst_cell.fill = _EMERALD_LIGHT

    # PI tracking columns (19-23): preserve existing values, never overwrite
    # Only set styling if cells are empty (new row)
    _PI_COL_START = 19  # ImageTrend Status
    _PI_COL_END = 23    # Last PI Update
    for col in range(_PI_COL_START, _PI_COL_END + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            cell.value = ""  # Initialize empty, not None
        cell.font = _BODY_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 22))


def _build_vrc_sheet(wb: Workbook, evaluation: Dict, vrc_results: List[Dict], row: int) -> None:
    """Populate VRC Injury Categories sheet."""
    ws = wb["VRC Injury Categories"]

    ws.cell(row=row, column=1, value=evaluation.get("patient_name", "")).font = _BOLD_FONT
    ws.cell(row=row, column=2, value=evaluation.get("patient_id", "")).font = _BODY_FONT

    vrc_map = {r["category_id"]: r for r in vrc_results}
    for i, cat_id in enumerate(_VRC_IDS):
        col = i + 3  # offset by name + ID columns
        r = vrc_map.get(cat_id, {})
        status = r.get("status", "UNABLE")
        cell = ws.cell(row=row, column=col, value=status)
        cell.font = _BODY_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

        # Color cells
        if status == "YES":
            cell.fill = _PINK_LIGHT
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFBE185D")
        elif status == "POSSIBLE":
            cell.fill = _GOLD_LIGHT
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFB8860B")
        elif status == "NO":
            cell.fill = _EMERALD_LIGHT
        else:
            cell.fill = _GRAY_FILL


def _build_protocol_sheet(wb: Workbook, evaluation: Dict, row: int) -> None:
    """Populate Protocol Detail sheet."""
    ws = wb["Protocol Detail"]

    ws.cell(row=row, column=1, value=evaluation.get("patient_name", "")).font = _BOLD_FONT
    ws.cell(row=row, column=2, value=evaluation.get("patient_id", "")).font = _BODY_FONT

    results = evaluation.get("results", [])
    triggered = [r for r in results if r["outcome"] != "NOT_TRIGGERED"]
    for i, r in enumerate(triggered):
        col = i + 3
        outcome = r.get("outcome", "")
        short = {"COMPLIANT": "C", "NON_COMPLIANT": "NC", "INDETERMINATE": "I", "ERROR": "E"}.get(outcome, outcome)
        cell = ws.cell(row=row, column=col, value=short)
        cell.font = _BODY_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

        if outcome == "NON_COMPLIANT":
            cell.fill = _PINK_LIGHT
        elif outcome == "INDETERMINATE":
            cell.fill = _GOLD_LIGHT
        elif outcome == "COMPLIANT":
            cell.fill = _EMERALD_LIGHT

    # Write protocol names in header if row == 2
    if row == 2:
        for i, r in enumerate(triggered):
            col = i + 3
            hdr = ws.cell(row=1, column=col, value=r.get("protocol_name", ""))
            hdr.font = _HEADER_FONT
            hdr.fill = _HEADER_FILL
            hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _build_ntds_sheet(wb: Workbook, evaluation: Dict, row: int) -> None:
    """Populate NTDS Events sheet."""
    ws = wb["NTDS Events"]

    ws.cell(row=row, column=1, value=evaluation.get("patient_name", "")).font = _BOLD_FONT
    ws.cell(row=row, column=2, value=evaluation.get("patient_id", "")).font = _BODY_FONT

    ntds = evaluation.get("ntds_results", [])
    for i, r in enumerate(ntds):
        col = i + 3
        outcome = r.get("outcome", "")
        cell = ws.cell(row=row, column=col, value=outcome)
        cell.font = _BODY_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

        if outcome == "YES":
            cell.fill = _PINK_LIGHT
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFBE185D")
        elif outcome == "UNABLE_TO_DETERMINE":
            cell.fill = _GOLD_LIGHT
        elif outcome == "NO":
            cell.fill = _EMERALD_LIGHT
        elif outcome == "EXCLUDED":
            cell.fill = _GRAY_FILL

    # Write event names in header if row == 2
    if row == 2:
        for i, r in enumerate(ntds):
            col = i + 3
            hdr = ws.cell(row=1, column=col, value=f"#{r.get('event_id', 0):02d} {r.get('canonical_name', '')}")
            hdr.font = _HEADER_FONT
            hdr.fill = _HEADER_FILL
            hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Create workbook
# ---------------------------------------------------------------------------

def _create_workbook() -> Workbook:
    """Create a new workbook with all 4 sheets and header rows."""
    wb = Workbook()

    # Sheet 1: Patient Dashboard
    ws1 = wb.active
    ws1.title = "Patient Dashboard"
    headers = [
        "Patient Name", "MRN / ID", "DOB", "Arrival", "Trauma Cat", "Status",
        "Triggered", "Compliant", "Non-Compliant", "Indeterminate",
        "NTDS YES", "NTDS YES Events", "NTDS Unable",
        "VRC YES", "VRC Possible", "Worst Outcome", "Report Link", "Last Evaluated",
        # PI tracking columns (operator-only — never auto-populated by engine)
        "ImageTrend Status", "PI Review Status", "PI Committee Date",
        "Reviewer Notes", "Last PI Update",
    ]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header_row(ws1, len(headers))
    ws1.freeze_panes = "C2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Column widths
    widths = [22, 12, 12, 20, 8, 12, 10, 10, 13, 13, 10, 35, 10, 30, 30, 16, 25, 20,
              16, 16, 16, 30, 20]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Data validation dropdowns for PI tracking columns
    dv_imagetrend = DataValidation(
        type="list", formula1='"Not Started,In Progress,Complete"', allow_blank=True
    )
    dv_imagetrend.prompt = "Select ImageTrend status"
    dv_imagetrend.promptTitle = "ImageTrend Status"
    ws1.add_data_validation(dv_imagetrend)
    dv_imagetrend.add(f"S2:S500")  # Column 19

    dv_pi_review = DataValidation(
        type="list", formula1='"Pending,In Review,Reviewed,Presented"', allow_blank=True
    )
    dv_pi_review.prompt = "Select PI review status"
    dv_pi_review.promptTitle = "PI Review Status"
    ws1.add_data_validation(dv_pi_review)
    dv_pi_review.add(f"T2:T500")  # Column 20

    # Sheet 2: VRC Injury Categories
    ws2 = wb.create_sheet("VRC Injury Categories")
    vrc_headers = ["Patient Name", "MRN / ID"] + [cid.replace("_", " ").title() for cid in _VRC_IDS]
    for col, h in enumerate(vrc_headers, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header_row(ws2, len(vrc_headers))
    ws2.freeze_panes = "C2"
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    for i in range(3, len(vrc_headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 14

    # Sheet 3: Protocol Detail
    ws3 = wb.create_sheet("Protocol Detail")
    ws3.cell(row=1, column=1, value="Patient Name").font = _HEADER_FONT
    ws3.cell(row=1, column=1).fill = _HEADER_FILL
    ws3.cell(row=1, column=2, value="MRN / ID").font = _HEADER_FONT
    ws3.cell(row=1, column=2).fill = _HEADER_FILL
    ws3.freeze_panes = "C2"
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 12

    # Sheet 4: NTDS Events
    ws4 = wb.create_sheet("NTDS Events")
    ws4.cell(row=1, column=1, value="Patient Name").font = _HEADER_FONT
    ws4.cell(row=1, column=1).fill = _HEADER_FILL
    ws4.cell(row=1, column=2, value="MRN / ID").font = _HEADER_FONT
    ws4.cell(row=1, column=2).fill = _HEADER_FILL
    ws4.freeze_panes = "C2"
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 12

    return wb


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def update_excel_dashboard(
    evaluation: Dict[str, Any],
    vrc_results: List[Dict[str, Any]],
    output_path: Path,
) -> Path:
    """
    Add or update a patient row in the Excel trauma dashboard.

    If the file exists, opens it and updates/appends.
    If not, creates a new workbook with all sheets and formatting.

    Args:
        evaluation: Patient evaluation dict from batch_eval.evaluate_patient()
        vrc_results: VRC classification results from classify_vrc_categories()
        output_path: Path to Excel file

    Returns:
        Path to the Excel file
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel dashboard generation")

    # Load or create workbook
    if output_path.exists():
        wb = load_workbook(output_path)
    else:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = _create_workbook()

    patient_id = str(evaluation.get("patient_id", ""))

    # Update/append to each sheet
    for sheet_name, builder in [
        ("Patient Dashboard", lambda row: _build_dashboard_sheet(wb, evaluation, vrc_results, row)),
        ("VRC Injury Categories", lambda row: _build_vrc_sheet(wb, evaluation, vrc_results, row)),
        ("Protocol Detail", lambda row: _build_protocol_sheet(wb, evaluation, row)),
        ("NTDS Events", lambda row: _build_ntds_sheet(wb, evaluation, row)),
    ]:
        ws = wb[sheet_name]
        existing_row = _find_patient_row(ws, patient_id)
        if existing_row:
            row = existing_row
        else:
            row = ws.max_row + 1
            if ws.max_row == 1 and ws.cell(row=1, column=2).value and not ws.cell(row=2, column=1).value:
                row = 2  # First data row
        builder(row)

    wb.save(output_path)
    return output_path
